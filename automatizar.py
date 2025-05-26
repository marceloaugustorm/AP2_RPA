import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Aluno e data
nome_aluno = "Carlos Eduardo RA:2400852 e Vinicius Gama: 2400375"
data_hoje = datetime.now().strftime("%d/%m/%Y")

# Novo estilo moderno
titulo_font = Font(name="Calibri", size=13, bold=True, color="000000")
titulo_fill = PatternFill("solid", fgColor="BDD7EE")
titulo_alignment = Alignment(horizontal="center", vertical="center")

cabecalho_font = Font(name="Calibri", bold=True, color="FFFFFF")
cabecalho_fill = PatternFill("solid", fgColor="4472C4")
cabecalho_alignment = Alignment(horizontal="center", vertical="center")

borda = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF")
)

alinhamento_padrao = Alignment(horizontal="center", vertical="center")
fill_zebra = PatternFill("solid", fgColor="F2F2F2")  # para zebra

def aplicar_estilo(ws, linha_inicio):
    for row in ws.iter_rows(min_row=linha_inicio, max_row=linha_inicio):
        for cell in row:
            cell.font = cabecalho_font
            cell.fill = cabecalho_fill
            cell.alignment = cabecalho_alignment
            cell.border = borda

    for i, row in enumerate(ws.iter_rows(min_row=linha_inicio + 1), start=0):
        for cell in row:
            cell.border = borda
            cell.alignment = alinhamento_padrao
            if i % 2 == 0:
                cell.fill = fill_zebra

def ajustar_largura(ws):
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 3

# Criar planilha
wb = Workbook()
ws1 = wb.active
ws1.title = "Países"

# Dados países
conn_paises = sqlite3.connect("paises.db")
cursor_paises = conn_paises.cursor()
cursor_paises.execute("SELECT * FROM countries")
dados_paises = cursor_paises.fetchall()
colunas_paises = [desc[0] for desc in cursor_paises.description]
conn_paises.close()

# Título da aba países
ws1.merge_cells('A1:I1')
ws1['A1'] = f"Relatório de Países – Aluno: {nome_aluno} – Data: {data_hoje}"
ws1['A1'].font = titulo_font
ws1['A1'].fill = titulo_fill
ws1['A1'].alignment = titulo_alignment

# Adicionar cabeçalhos e dados
linha_inicio_paises = 3
ws1.append([])
ws1.append(colunas_paises)

for linha in dados_paises:
    ws1.append(linha)

aplicar_estilo(ws1, linha_inicio_paises)
ajustar_largura(ws1)

# Aba livros
ws2 = wb.create_sheet("Livros")
conn_livros = sqlite3.connect("livraria.db")
cursor_livros = conn_livros.cursor()
cursor_livros.execute("SELECT * FROM livros")
dados_livros = cursor_livros.fetchall()
colunas_livros = [desc[0] for desc in cursor_livros.description]
conn_livros.close()

# Título da aba livros
ws2.merge_cells('A1:E1')
ws2['A1'] = f"Relatório de Livros – Aluno: {nome_aluno} – Data: {data_hoje}"
ws2['A1'].font = titulo_font
ws2['A1'].fill = titulo_fill
ws2['A1'].alignment = titulo_alignment

linha_inicio_livros = 3
ws2.append([])
ws2.append(colunas_livros)

for linha in dados_livros:
    ws2.append(linha)

aplicar_estilo(ws2, linha_inicio_livros)
ajustar_largura(ws2)

# Salvar
wb.save("relatorio_moderno.xlsx")
print("Novo relatório salvo como: relatorio_moderno.xlsx")
